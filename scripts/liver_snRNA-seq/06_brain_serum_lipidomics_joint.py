"""
脑-血清联合脂质组分析
比较 5dpi 时脑组织与血清脂质的变化方向，寻找摄取/消耗模式

数据：
  脑: data/lipidomics-5dpi/brain/brain-all-lipids.numbers
  血清: data/lipidomics-5dpi/serum/serum-all-lipids-5dpi_vs_control.xlsx
"""

import shutil, re
import pandas as pd
import numpy as np
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import os

# numbers-parser 对含空格/特殊字符的路径有 bug，先复制到 /tmp
BRAIN_SRC  = '/Users/maxue/Documents/vscode/tbi/data/lipidomics-5dpi/brain/brain-all-lipids.numbers'
BRAIN_TMP  = '/tmp/brain_all_lipidomics.numbers'
SERUM_PATH = '/Users/maxue/Documents/vscode/tbi/data/lipidomics-5dpi/serum/serum-all-lipids-5dpi_vs_control.xlsx'

FIG_DIR = '/Users/maxue/Documents/vscode/tbi/results/lipidomics_joint/figures'
TAB_DIR = '/Users/maxue/Documents/vscode/tbi/results/lipidomics_joint/tables'
os.makedirs(FIG_DIR, exist_ok=True)
os.makedirs(TAB_DIR, exist_ok=True)

shutil.copy(BRAIN_SRC, BRAIN_TMP)

# ── 读取脑组织脂质组 ──────────────────────────────────────────────
from numbers_parser import Document
doc   = Document(BRAIN_TMP)
table = doc.sheets[0].tables[0]
data  = [[c.value for c in row] for row in table.iter_rows()]
brain = pd.DataFrame(data[1:], columns=data[0])
brain.columns = ['Index','Formula','Compounds','cn_name','FC','Log2FC','Type',
                 'ClassI','cn_classI','ClassII','cn_classII','lipidmaps',
                 'swi1','swi2','swi3','swi4','swi5','swi6','swi7',
                 'CON1','CON2','CON3','CON4','CON5','CON6','CON7','VIP','pval','FDR']
for c in ['FC','Log2FC','pval','FDR']: brain[c] = pd.to_numeric(brain[c], errors='coerce')

# ── 读取血清脂质组 ─────────────────────────────────────────────────
wb    = openpyxl.load_workbook(SERUM_PATH)
rows  = list(wb.active.iter_rows(values_only=True))
serum = pd.DataFrame(rows[1:], columns=list(rows[0]))
serum.columns = ['Index','Formula','Compounds','cn_name','ClassI','cn_classI',
                 'ClassII','cn_classII','lipidmaps','FC','Type','pval',
                 's1','s2','s3','s6','c3','c4','c5','c6','VIP','Log2FC','FDR']
for c in ['FC','Log2FC','pval','FDR']: serum[c] = pd.to_numeric(serum[c], errors='coerce')

print(f"脑组织: {len(brain)} 种脂质  up={(brain['Type']=='up').sum()}  down={(brain['Type']=='down').sum()}")
print(f"血清:   {len(serum)} 种脂质  up={(serum['Type']=='up').sum()}  down={(serum['Type']=='down').sum()}")

# ── 1. 各亚类显著变化汇总 ─────────────────────────────────────────
brain_sig = brain[brain['Type'].isin(['up','down'])]
sub = brain_sig.groupby(['ClassII','Type']).size().unstack(fill_value=0)
sub['total'] = sub.sum(axis=1)
sub.sort_values('total', ascending=False).to_csv(f'{TAB_DIR}/brain_sig_by_class.csv')

# ── 2. 溶血磷脂方向对比（脑 vs 血清）────────────────────────────
lyso_summary = []
for tissue, df in [('brain', brain), ('serum', serum)]:
    lyso = df[df['ClassII'].str.match(r'^L(PC|PE|PA|PS|PI|PG)', na=False)]
    for cl in lyso['ClassII'].unique():
        sub_cl = lyso[lyso['ClassII'] == cl]
        lyso_summary.append({
            'tissue': tissue, 'ClassII': cl,
            'n': len(sub_cl),
            'n_up':   (sub_cl['Type']=='up').sum(),
            'n_down': (sub_cl['Type']=='down').sum(),
            'mean_log2FC': sub_cl['Log2FC'].mean(),
        })
lyso_df = pd.DataFrame(lyso_summary)
lyso_df.to_csv(f'{TAB_DIR}/lyso_direction_brain_vs_serum.csv', index=False)
print("\n溶血磷脂方向汇总（脑 vs 血清）：")
print(lyso_df.pivot_table(index='ClassII', columns='tissue',
      values=['n_up','n_down','mean_log2FC']).to_string())

# ── 3. PUFA 磷脂方向对比 ─────────────────────────────────────────
pufa_rows = []
for label, kw in [('DHA(22:6)','22:6'), ('AA(20:4)','20:4'), ('EPA(20:5)','20:5')]:
    for tissue, df in [('brain',brain), ('serum',serum)]:
        sub = df[(df['ClassI']=='GP') & df['Compounds'].str.contains(kw, na=False)]
        pufa_rows.append({
            'PUFA': label, 'tissue': tissue, 'n': len(sub),
            'n_up':   (sub['Type']=='up').sum(),
            'n_down': (sub['Type']=='down').sum(),
            'mean_log2FC': sub['Log2FC'].mean(),
        })
pufa_df = pd.DataFrame(pufa_rows)
pufa_df.to_csv(f'{TAB_DIR}/pufa_phospholipid_direction.csv', index=False)
print("\nPUFA 磷脂方向（脑 vs 血清）：")
print(pufa_df.to_string(index=False))

# ── 4. TG 不饱和度分组对比 ───────────────────────────────────────
def tg_group(df, label):
    tg = df[df['ClassII']=='TG'].copy()
    tg['db'] = tg['Compounds'].apply(
        lambda s: sum(int(x) for x in re.findall(r':(\d+)', str(s))))
    tg['grp'] = pd.cut(tg['db'], bins=[-1,2,5,8,100],
                       labels=['≤2 sat/MUFA','3-5 低PUFA','6-8 含DHA','≥9 高PUFA'])
    g = tg.groupby('grp', observed=True).agg(
        mean_log2FC=('Log2FC','mean'), n=('Log2FC','count'),
        n_up=('Type', lambda x: (x=='up').sum()),
        n_down=('Type', lambda x: (x=='down').sum()))
    g.columns = [f'{label}_{c}' for c in g.columns]
    return g

tg_cmp = pd.concat([tg_group(brain,'脑'), tg_group(serum,'血清')], axis=1)
tg_cmp.to_csv(f'{TAB_DIR}/TG_unsaturation_brain_vs_serum.csv')
print("\nTG 不饱和度分组对比：")
print(tg_cmp.to_string())

# ── 5. 脑↑ 且 血清↓ 的脂质（摄取信号）───────────────────────────
b_up = set(brain[brain['Type']=='up']['Compounds'])
s_dn = set(serum[serum['Type']=='down']['Compounds'])
overlap_up = b_up & s_dn
ov_b = brain[brain['Compounds'].isin(overlap_up)][['Compounds','ClassII','Log2FC']].rename(columns={'Log2FC':'脑_log2FC'})
ov_s = serum[serum['Compounds'].isin(overlap_up)][['Compounds','Log2FC']].rename(columns={'Log2FC':'血清_log2FC'})
uptake_df = ov_b.merge(ov_s, on='Compounds').sort_values('脑_log2FC', ascending=False)
uptake_df.to_csv(f'{TAB_DIR}/brain_up_serum_down_uptake.csv', index=False)
print(f"\n脑↑ 且 血清↓（摄取信号）：{len(uptake_df)} 种")
print(uptake_df.to_string(index=False))

# ── 6. 脑显著上调的 DHA/AA 脂质 ──────────────────────────────────
for kw, name in [('22:6','DHA'), ('20:4','AA')]:
    sub = brain[(brain['Type']=='up') & brain['Compounds'].str.contains(kw, na=False)]
    sub.sort_values('Log2FC', ascending=False)[
        ['Compounds','ClassII','FC','Log2FC','pval']].to_csv(
        f'{TAB_DIR}/brain_up_{name}_lipids.csv', index=False)

# ── 7. 可视化：溶血磷脂方向对比图 ────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(12, 5))
for ax, (tissue, df), color in zip(axes,
        [('脑组织（5dpi）', brain, '#e74c3c'),
         ('血清（5dpi）',   serum, '#2980b9')],
        ['#e74c3c','#2980b9']):
    lyso = df[df['ClassII'].str.match(r'^L(PC|PE|PA|PS|PI|PG)', na=False)]
    by_cl = lyso.groupby('ClassII')['Log2FC'].mean().sort_values(ascending=False)
    bars = ax.bar(range(len(by_cl)), by_cl.values, color=color, alpha=0.75)
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xticks(range(len(by_cl)))
    ax.set_xticklabels(by_cl.index, rotation=30, ha='right')
    ax.set_ylabel('Mean log2FC')
    ax.set_title(tissue)
plt.suptitle('溶血磷脂各亚类方向：脑组织 vs 血清（5dpi vs control）', fontsize=12)
plt.tight_layout()
plt.savefig(f'{FIG_DIR}/lyso_direction_brain_vs_serum.png', dpi=150, bbox_inches='tight')
plt.close()

# ── 8. 散点图：同一脂质在脑和血清中的 log2FC ────────────────────
common = brain[['Compounds','ClassII','Log2FC']].rename(columns={'Log2FC':'brain_log2FC'}).merge(
         serum[['Compounds','Log2FC']].rename(columns={'Log2FC':'serum_log2FC'}), on='Compounds')
common = common.dropna()

fig, ax = plt.subplots(figsize=(7, 7))
# 按亚类着色
classes = common['ClassII'].unique()
palette = dict(zip(classes, sns.color_palette('tab20', len(classes))))
for cl, grp in common.groupby('ClassII'):
    ax.scatter(grp['serum_log2FC'], grp['brain_log2FC'],
               s=10, alpha=0.5, label=cl, color=palette[cl])
ax.axhline(0, color='grey', linewidth=0.6, linestyle='--')
ax.axvline(0, color='grey', linewidth=0.6, linestyle='--')
ax.set_xlabel('血清 log2FC (5dpi/control)')
ax.set_ylabel('脑组织 log2FC (5dpi/control)')
ax.set_title('脑-血清脂质变化方向散点图\n（左上象限 = 脑↑血清↓ = 摄取信号）')
# 标注"脑↑血清↓"象限
ax.fill_betweenx([-3, 4], -3, 0, alpha=0.04, color='red')
ax.fill_between([-3, 0], 0, 4, alpha=0.04, color='red')
ax.text(-2.5, 3.5, '脑↑血清↓\n(摄取)', color='red', fontsize=10)
ax.set_xlim(-3, 3); ax.set_ylim(-4, 5)
leg = ax.legend(fontsize=5, ncol=3, loc='lower right', markerscale=2)
plt.tight_layout()
plt.savefig(f'{FIG_DIR}/brain_serum_log2FC_scatter.png', dpi=150, bbox_inches='tight')
plt.close()

print("\n=== 脑-血清联合脂质组分析完成 ===")
print(f"结果图表: {FIG_DIR}")
print(f"结果表格: {TAB_DIR}")
